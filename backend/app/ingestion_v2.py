import os
import re
import csv
import io
import math
import chardet
import openpyxl
import xlrd
import numpy as np
import pandas as pd
import duckdb
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional
from app.session_manager import Session

MISSING_VALUE_STRINGS = {
    "n/a", "na", "n/d", "nd", "-", "--", "none", "null", "nan", "#n/a", "#div/0!", "#ref!", "#value!", "#name?", "#num!"
}

TOTAL_KEYWORDS = {"total", "sous-total", "soustotal", "somme", "subtotal", "moyenne", "grand total"}

class IngestionResult:
    def __init__(
        self,
        table_name: str,
        confidence_score: float,
        requires_user_action: bool,
        detected_header_index: int,
        selected_sheet: Optional[str],
        available_sheets: List[str],
        warnings: List[str],
        raw_preview_rows: List[List[Any]]
    ):
        self.table_name = table_name
        self.confidence_score = confidence_score
        self.requires_user_action = requires_user_action
        self.detected_header_index = detected_header_index
        self.selected_sheet = selected_sheet
        self.available_sheets = available_sheets
        self.warnings = warnings
        self.raw_preview_rows = raw_preview_rows

def detect_file_format_and_magic_bytes(file_path: Path, filename: str) -> Tuple[str, List[str]]:
    warnings = []
    with open(file_path, "rb") as f:
        header_bytes = f.read(8)

    ext = Path(filename).suffix.lower()

    if header_bytes.startswith(b"PK\x03\x04"):
        actual_format = "xlsx"
        if ext not in [".xlsx", ".xlsm"]:
            warnings.append(f"Extension '{ext}' incohérente. Le fichier est un classeur Excel moderne (.xlsx).")
    elif header_bytes.startswith(b"\xd0\xcf\x11\xe0"):
        actual_format = "xls"
        if ext != ".xls":
            warnings.append(f"Extension '{ext}' incohérente. Le fichier est un ancien format Excel (.xls).")
    else:
        actual_format = "csv"
        if ext in [".xlsx", ".xls"]:
            warnings.append(f"Le fichier a une extension '{ext}' mais contient du texte brut (CSV/TSV).")

    return actual_format, warnings

def sniff_csv_delimiter_and_encoding(file_path: Path) -> Tuple[str, str, List[str]]:
    warnings = []
    with open(file_path, "rb") as f:
        raw_bytes = f.read(50000)

    detected = chardet.detect(raw_bytes)
    encoding = detected.get("encoding") or "utf-8"
    if encoding.lower() in ["ascii", "utf-8", "utf-8-sig"]:
        encoding = "utf-8"
    else:
        warnings.append(f"Encodage du fichier CSV détecté : '{encoding}'. Conversion automatique vers UTF-8.")

    try:
        sample_text = raw_bytes.decode(encoding, errors="ignore")
    except Exception:
        sample_text = raw_bytes.decode("latin-1", errors="ignore")
        encoding = "latin-1"

    candidates = [",", ";", "\t", "|"]
    counts = {c: sample_text.count(c) for c in candidates}
    delimiter = max(counts, key=counts.get) if max(counts.values(), default=0) > 0 else ","

    if delimiter == ";":
        warnings.append("Séparateur point-virgule (';') détecté et appliqué pour le CSV.")
    elif delimiter == "\t":
        warnings.append("Séparateur tabulation ('\\t') détecté et appliqué pour le CSV.")

    return encoding, delimiter, warnings

def read_raw_rows_csv(file_path: Path, encoding: str, delimiter: str, max_rows: int = 25) -> List[List[Any]]:
    raw_rows = []
    with open(file_path, "r", encoding=encoding, errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            raw_rows.append(row)
    return raw_rows

def read_excel_data(file_path: Path, actual_format: str, sheet_name: Optional[str] = None) -> Tuple[List[List[Any]], List[List[Any]], List[str], str]:
    """
    Reads Excel workbook, propagates merged cells, and returns (raw_preview_rows, all_unmerged_rows, available_sheets, selected_sheet)
    """
    available_sheets = []
    selected_sheet = ""
    all_rows = []

    if actual_format == "xlsx":
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as e:
            err_str = str(e).lower()
            if "protected" in err_str or "encrypted" in err_str or "password" in err_str:
                raise ValueError("Le fichier Excel est protégé par un mot de passe. Veuillez le déverrouiller avant importation.")
            raise ValueError(f"Fichier Excel corrompu ou illisible: {str(e)}")

        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state == "visible" and ws.max_row > 0 and ws.max_column > 0:
                available_sheets.append(name)

        if not available_sheets:
            available_sheets = wb.sheetnames

        selected_sheet = sheet_name if sheet_name in available_sheets else available_sheets[0]
        ws = wb[selected_sheet]

        # Unmerge cells and propagate top-left value across range
        merged_ranges = list(ws.merged_cells.ranges)
        for mrange in merged_ranges:
            top_left_value = ws.cell(mrange.min_row, mrange.min_col).value
            ws.unmerge_cells(str(mrange))
            for row in range(mrange.min_row, mrange.max_row + 1):
                for col in range(mrange.min_col, mrange.max_col + 1):
                    ws.cell(row, col).value = top_left_value

        for row in ws.iter_rows(values_only=True):
            all_rows.append([cell for cell in row])

    else:
        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as e:
            if "protected" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError("Le fichier Excel (.xls) est protégé par mot de passe.")
            raise ValueError(f"Fichier Excel (.xls) illisible ou corrompu: {str(e)}")

        available_sheets = wb.sheet_names()
        selected_sheet = sheet_name if sheet_name in available_sheets else available_sheets[0]
        ws = wb.sheet_by_name(selected_sheet)

        for r_idx in range(ws.nrows):
            all_rows.append(ws.row_values(r_idx))

    preview_rows = all_rows[:25]
    return preview_rows, all_rows, available_sheets, selected_sheet

def score_header_candidates(raw_rows: List[List[Any]]) -> Tuple[int, float]:
    if not raw_rows:
        return 0, 0.0

    best_idx = 0
    best_score = -1.0

    for idx, row in enumerate(raw_rows[:15]):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if not non_empty:
            continue

        row_len = len(row)
        if row_len == 0:
            continue

        non_empty_ratio = len(non_empty) / row_len

        # Penalize single-cell title rows
        if len(non_empty) == 1:
            non_empty_ratio *= 0.2

        string_count = sum(1 for c in non_empty if isinstance(c, str) and not c.strip().replace(".", "", 1).isdigit())
        string_ratio = string_count / len(non_empty) if non_empty else 0

        str_vals = [str(c).strip().lower() for c in non_empty]
        unique_ratio = len(set(str_vals)) / len(str_vals) if str_vals else 0
        len_score = sum(1 for c in str_vals if 2 <= len(c) <= 60) / len(str_vals) if str_vals else 0

        subsequent_rows = raw_rows[idx + 1: idx + 6]
        type_break_score = 0.5
        if subsequent_rows:
            num_counts = 0
            for s_row in subsequent_rows:
                for val in s_row:
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        num_counts += 1
            if num_counts > 0:
                type_break_score = 1.0

        score = (
            (non_empty_ratio * 0.35) +
            (string_ratio * 0.25) +
            (unique_ratio * 0.20) +
            (len_score * 0.10) +
            (type_break_score * 0.10)
        )

        if score > best_score:
            best_score = score
            best_idx = idx

    confidence = round(best_score, 2) if best_score > 0 else 0.0
    return best_idx, confidence

def sanitize_dataframe(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    warnings = []

    # 1. Trim phantom empty rows and columns
    df = df.dropna(how="all").dropna(how="all", axis=1)

    # 2. Column renaming & sanitization
    new_cols = []
    seen_cols = {}

    for idx, col in enumerate(df.columns):
        col_raw = str(col).strip() if col is not None else ""
        if pd.isna(col) or not col_raw or col_raw.lower() in ["none", "nan"] or col_raw.lower().startswith("unnamed"):
            col_str = f"Colonne_{idx + 1}"
            warnings.append(f"Colonne sans nom détectée et renommée '{col_str}'.")
        else:
            col_str = col_raw

        # Sanitize special characters for DuckDB SQL (keep alphanumeric, accented chars and underscores)
        col_str = re.sub(r'[^\w\s_]', '_', col_str)
        col_str = re.sub(r'\s+', '_', col_str)

        if col_str in seen_cols:
            seen_cols[col_str] += 1
            col_sanitized = f"{col_str}_{seen_cols[col_str]}"
            warnings.append(f"Colonne dupliquée '{col_str}' automatiquement renommée en '{col_sanitized}'.")
        else:
            seen_cols[col_str] = 0
            col_sanitized = col_str

        new_cols.append(col_sanitized)

    df.columns = new_cols

    # 3. Clean intercalated Total / Subtotal rows (vectorized)
    total_pattern = re.compile(r'\b(?:' + '|'.join(re.escape(kw) for kw in TOTAL_KEYWORDS) + r')\b', re.IGNORECASE)
    mask_total = df.astype(str).apply(lambda s: s.str.contains(total_pattern, na=False)).any(axis=1)
    if mask_total.any():
        num_totals = int(mask_total.sum())
        df = df[~mask_total]
        warnings.append(f"{num_totals} ligne(s) de total/sous-total détectée(s) et ignorée(s) pour l'analyse.")

    # 4. Column Data Normalization
    for col in df.columns:
        # Convert missing value strings and Excel errors
        df[col] = df[col].apply(
            lambda v: None if pd.isna(v) or v is None or str(v).strip().lower() in MISSING_VALUE_STRINGS or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))) else v
        )

        non_nulls = df[col].dropna()
        if non_nulls.empty:
            continue

        first_val = non_nulls.iloc[0]

        # Excel Serial Date Conversion (numbers between 35000 and 55000)
        if isinstance(first_val, (int, float, np.integer, np.floating)) and not isinstance(first_val, bool):
            if ("date" in col.lower() or "embauche" in col.lower()) or (35000 <= first_val <= 55000):
                try:
                    df[col] = df[col].apply(
                        lambda v: (datetime(1899, 12, 30) + timedelta(days=float(v))).strftime("%Y-%m-%d") if isinstance(v, (int, float, np.integer, np.floating)) and 35000 <= float(v) <= 55000 else v
                    )
                    warnings.append(f"Colonne '{col}' : dates en format numérique Excel automatiquement converties en format ISO (AAAA-MM-JJ).")
                except Exception:
                    pass

        # French regional numbers ('1 234,56' -> 1234.56)
        if isinstance(first_val, str):
            french_num_pattern = re.compile(r'^-?\d{1,3}(\s?\d{3})*(,\d+)?$')
            if all(bool(french_num_pattern.match(str(v).strip())) for v in non_nulls[:20]):
                try:
                    df[col] = df[col].apply(
                        lambda v: float(str(v).replace(" ", "").replace(",", ".")) if v is not None and str(v).strip() != "" else None
                    )
                    warnings.append(f"Colonne '{col}' : nombres au format régional français (ex: '1 234,56') convertis en numériques.")
                except Exception:
                    pass

    return df, warnings

def ingest_file_v2(
    session: Session,
    file_path: Path,
    filename: str,
    selected_sheet: Optional[str] = None,
    header_index: Optional[int] = None,
    delimiter: Optional[str] = None
) -> IngestionResult:
    all_warnings = []

    actual_format, format_warns = detect_file_format_and_magic_bytes(file_path, filename)
    all_warnings.extend(format_warns)

    encoding = "utf-8"
    if actual_format == "csv":
        enc, sn_delim, csv_warns = sniff_csv_delimiter_and_encoding(file_path)
        all_warnings.extend(csv_warns)
        encoding = enc
        chosen_delimiter = delimiter or sn_delim
        raw_rows = read_raw_rows_csv(file_path, encoding, chosen_delimiter, max_rows=25)
        available_sheets = []
        sheet_name = None
    else:
        raw_rows, all_unmerged_rows, available_sheets, sheet_name = read_excel_data(file_path, actual_format, selected_sheet)
        chosen_delimiter = None
        if len(available_sheets) > 1 and not selected_sheet:
            all_warnings.append(f"Classeur multi-feuilles détecté ({len(available_sheets)} feuilles). La feuille '{sheet_name}' a été sélectionnée par défaut.")

    if header_index is not None:
        best_header_idx = header_index
        confidence_score = 1.0
    else:
        best_header_idx, confidence_score = score_header_candidates(raw_rows)

    requires_user_action = False
    if confidence_score < 0.75 and header_index is None:
        requires_user_action = True
        all_warnings.append("Confiance faible sur la détection de la ligne d'en-tête. Veuillez vérifier la sélection dans l'aperçu brut.")

    if len(available_sheets) > 1 and not selected_sheet:
        requires_user_action = True

    if actual_format == "csv":
        df_raw = pd.read_csv(
            file_path,
            encoding=encoding,
            sep=chosen_delimiter,
            header=best_header_idx,
            on_bad_lines="skip"
        )
    else:
        if all_unmerged_rows and best_header_idx < len(all_unmerged_rows):
            header_row = all_unmerged_rows[best_header_idx]
            data_rows = all_unmerged_rows[best_header_idx + 1:]
            df_raw = pd.DataFrame(data_rows, columns=header_row)
        else:
            df_raw = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=best_header_idx
            )

    clean_df, clean_warns = sanitize_dataframe(df_raw)
    all_warnings.extend(clean_warns)

    conn = session.get_connection()
    table_name = "dataset"
    conn.execute(f"DROP TABLE IF EXISTS {table_name}")

    conn.register("temp_clean_df", clean_df)
    conn.execute(f"CREATE TABLE {table_name} AS SELECT * FROM temp_clean_df")
    conn.unregister("temp_clean_df")

    session.dataset_filename = filename
    session.table_name = table_name

    raw_preview = [[str(cell) if cell is not None else "" for cell in r] for r in raw_rows[:20]]

    return IngestionResult(
        table_name=table_name,
        confidence_score=confidence_score,
        requires_user_action=requires_user_action,
        detected_header_index=best_header_idx,
        selected_sheet=sheet_name,
        available_sheets=available_sheets,
        warnings=list(dict.fromkeys(all_warnings)),
        raw_preview_rows=raw_preview
    )
